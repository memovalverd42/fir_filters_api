"""
Autor: Valverde Baez Guillermo

Procesamiento Digital de Señales

Instituto Tecnologico de Veracruz
"""

import math
import numpy as np
from scipy.signal import freqz
import matplotlib.pyplot as plt
from openpyxl import Workbook

""" DATOS DEL PROBLEMA Y CONFIGURACIÓN """

#### LOWPASS O HIGHPASS #####
Ap = 0.4  # 0.3       # Rizo en la banda pasante
As = 34  # 35      # Limite de atenuacion
fp = 16  # 50000   # Frecuencia del polo
fs = 8  # 40000   # Frecuencia del cero
F = 80  # 3000   # Frecuencia de muestreo

#### BANDPASS O STOPBAND ######
fs1 = 200  # Frecuencia del cero 1
fp1 = 100  # Frecuencia del polo 1
fs2 = 400  # Frecuencia del cero 2
fp2 = 700  # Frecuencia del polo 2

#### TIPO DE FILTRO -- HP = HIGHPASS | LP = LOWPASS | BP = BANDPASS | BS = STOPBAND   ####
filter_type = 'HP'
#### TIPO DE VENTANA -- H = HAMMING | B = BLACKMANN | K = KAISER   ####
window = 'H'

"""################################################ DELTA #########################################################"""

delta = 0
delta_s = 10 ** (-0.05 * As)
delta_p = (10 ** (0.05 * Ap) - 1) / (10 ** (0.05 * Ap) + 1)
delta = round(min(delta_s, delta_p), 7)

"""################################################## As y Ap ######################################################"""

AS = round(-20 * math.log10(delta), 7)
AP = round(20 * math.log10((1 + delta) / (1 - delta)), 7)

"""############################################# PARAMETRO D ######################################################"""

D = 0
if AS <= 21:
    D = 0.9222
else:
    D = round((AS - 7.95) / (14.36), 7)

"""############################################ PARAMETRO ALPHA ################################################"""

alpha = 0
if AS <= 21:
    alpha = 0
elif AS > 21 and AS <= 50:
    alpha = round((0.5842 * (AS - 21) ** (0.4)) + 0.07886 * (AS - 21), 7)
else:
    alpha = round(0.1102 * (AS - 8.7), 7)

"""############################################## ORDEN DEL FILTRO ###################################################"""
if filter_type == 'HP':
    N = int((((F) * (D)) / (fp - fs)) + 2)
elif filter_type == 'LP':
    N = int((((F) * (D)) / (fs - fp)) + 2)
elif filter_type == 'BP':
    N = round(((F * D) / (min(fp1 - fs1, fs2 - fp2))) + 1, 7)
elif filter_type == 'BS':
    N = round(((F * D) / (min(fs1 - fp1, fp2 - fs2))) + 1, 7)

N_o = N

N_int = int(N)
if (N_int + 1) % 2 == 0:
    N = N_int + 2
else:
    N = N_int + 1

n = int((N - 1) / 2)

print(f'\n delta = {delta} | As = {AS} | D = {D} | alpha = {alpha} | N = {N} | n = {n} | No ={N_o}\n')

"""###################################### TIPOS DE FILTROS #######################################################"""


def lowpass():  # FILTRO PASO BAJO
    coef = []
    nc = 1
    fc = 0.5 * (fp + fs)
    n0 = (2 * fc) / F
    coef.append(n0)
    while (nc <= n):
        term = (2 * math.pi * nc * fc) / F
        c = n0 * ((math.sin(term)) / (term))
        nc = nc + 1
        coef.append(round(c, 7))

    return coef


def highpass():  # FILTRO PASO ALTO
    coef = []
    nc = 1
    fc = 0.5 * (fp + fs)
    n0 = 1 - ((2 * fc) / F)
    coef.append(n0)
    while (nc <= n):
        term = (2 * math.pi * nc * fc) / F
        c = -((2 * fc) / F) * ((math.sin(term)) / (term))
        nc = nc + 1
        coef.append(round(c, 7))

    return coef


def bandpass():  # FILTRO PASO DE BANDA
    coef = []
    nc = 1

    deltaF = min(fp1 - fs1, fs2 - fp2)
    fc1 = fp1 - (deltaF / 2)
    fc2 = fp2 + (deltaF / 2)

    n0 = (2 / F) * (fc2 - fc1)
    coef.append(n0)

    while (nc <= n):
        term1 = (2 * math.pi * nc * fc2) / F
        term2 = (2 * math.pi * nc * fc1) / F
        c = (1 / (nc * math.pi)) * (math.sin(term1) - math.sin(term2))
        nc = nc + 1
        coef.append(round(c, 7))

    return coef


def bandstop():  # FILTRO RECHAZO DE BANDA
    coef = []
    nc = 1

    deltaF = min(fs1 - fp1, fp2 - fs2)
    fc1 = fp1 + (deltaF / 2)
    fc2 = fp2 - (deltaF / 2)

    n0 = (2 / F) * (fc1 - fc2) + 1
    coef.append(n0)

    while (nc <= n):
        term1 = (2 * math.pi * nc * fc1) / F
        term2 = (2 * math.pi * nc * fc2) / F
        c = (1 / (nc * math.pi)) * ((math.sin(term1)) - (math.sin(term2)))
        nc = nc + 1
        coef.append(round(c, 7))

    return coef


"""###################################### VENTANAS #######################################################"""


####################### VENTANA DE KAISER ####################
def sumatoria(beta):
    k = 1
    result = 0
    while (k <= 25):
        sum = ((1 / math.factorial(k)) * (beta / 2) ** k) ** 2
        result = result + round(sum, 7)
        k = k + 1

    return result + 1


def beta():
    betas = []
    nc = 0

    while (nc <= n):
        b = alpha * (1 - ((2 * nc) / (N - 1)) ** 2) ** 0.5
        betas.append(round(b, 7))
        nc = nc + 1

    return betas


betas = beta()


def kaiser():
    coef_k = []

    I_alpha = sumatoria(alpha)
    for beta in betas:
        coef = sumatoria(beta) / I_alpha
        coef_k.append(round(coef, 7))

    return coef_k


####################### VENTANA DE HAMMING ####################

def hamming():
    nc = 0
    coef = []
    while (nc <= n):
        r = 0.54 + 0.46 * math.cos((2 * math.pi * nc) / (N - 1))
        nc = nc + 1
        coef.append(round(r, 6))
    return coef


####################### VENTANA DE BLACKMANN ####################

def blackman():
    nc = 0
    coef = []
    while (nc <= n):
        r = 0.42 + 0.5 * math.cos((2 * math.pi * nc) / (N - 1)) + 0.08 * math.cos((4 * math.pi * nc) / (N - 1))
        nc = nc + 1
        coef.append(round(r, 6))

    return coef


"""###################################### FUNCION PARA ORDENAR COEFICIENTES #######################################################"""


def ordenamela(coefis):
    idx = len(coefis) - 1
    cont = 1
    coef_ordenados = []
    while (idx >= 0):
        if coefis[idx] == -0.0 or coefis[idx] == 0.0:
            cero = int(np.abs(coefis[idx]))
            coef_ordenados.append(cero)
        else:
            coef_ordenados.append(coefis[idx])
        idx = idx - 1

    while (cont <= len(coefis) - 1):
        if coefis[cont] == -0.0 or coefis[cont] == 0.0:
            cero = int(np.abs(coefis[cont]))
            coef_ordenados.append(cero)
        else:
            coef_ordenados.append(coefis[cont])
        cont = cont + 1

    return coef_ordenados


"""###################################### EJECUCIÓN #######################################################"""

if filter_type == 'HP':
    hd = highpass()
elif filter_type == 'LP':
    hd = lowpass()
elif filter_type == 'BP':
    hd = bandpass()
elif filter_type == 'BS':
    hd = bandstop()

if window == 'H':
    coef_v = hamming()
elif window == 'B':
    coef_v = blackman()
elif window == 'K':
    coef_v = kaiser()

coef_filt = []

for i in range(n + 1):
    coef_filt.append(round(coef_v[i] * hd[i], 7))

coef_filt_ordenados = ordenamela(coef_filt)

print('Ventana: \n')
print(coef_v, '\n')
print('Coeficientes: \n')
print(hd, '\n')
print('Coeficientes del Filtro: \n')
print(coef_filt, '\n')
print('Coeficientes ordenados: \n')
print(coef_filt_ordenados, '\n')
print('\n')

"""###################################### EXPORTAR DATOS A EXCEL #######################################################"""

wb = Workbook()
ruta = 'ventana.xlsx'

hoja = wb.active

fila = 2  # Fila donde empezamos

for val, dato, h, coeficiente in zip(range(n + 1), coef_v, hd, coef_filt):
    hoja.cell(column=2, row=fila, value=val)
    hoja.cell(column=3, row=fila, value=dato)
    hoja.cell(column=4, row=fila, value=h)
    hoja.cell(column=5, row=fila, value=coeficiente)
    fila += 1

wb.save(filename=ruta)

"""###################################### GRAFICACION #######################################################"""

num = coef_filt_ordenados
den = 1

w, h = freqz(num, den, 100)

hertz = w * F / (2 * np.pi)

plt.semilogy(hertz, np.abs(h))
plt.grid()
plt.show()



